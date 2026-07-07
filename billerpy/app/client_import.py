from io import BytesIO
from typing import Any

from openpyxl import load_workbook

COLUMN_MAP = {
    "tan": "tan",
    "pan": "pan",
    "company name": "name",
    "person name": "person_name",
    "gstn no": "gstin",
    "gstin": "gstin",
    "gstn": "gstin",
    "phone no": "phone",
    "pr phone no": "pr_phone",
    "pr mobile no": "pr_mobile",
}

SKIP_HEADERS = {
    "srl no",
    "sr no",
    "serial no",
    "email",
    "ain no",
    "ain",
    "category",
}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().replace(".", "").split())


def _cell_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _find_header_row(rows: list[tuple]) -> tuple[int, tuple] | None:
    for index, row in enumerate(rows):
        for header in row:
            normalized = _normalize_header(header)
            if normalized in COLUMN_MAP:
                return index, row
    return None


def parse_clients_xlsx(file_bytes: bytes) -> tuple[list[dict[str, str | None]], list[str]]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    all_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not all_rows:
        return [], ["The spreadsheet is empty."]

    header_match = _find_header_row(all_rows[:30])
    if not header_match:
        return [], ["Could not find a 'Company Name' column in the spreadsheet."]

    header_index, header_row = header_match
    data_rows = all_rows[header_index + 1:]

    field_indexes: dict[str, int] = {}
    for index, header in enumerate(header_row):
        normalized = _normalize_header(header)
        if not normalized or normalized in SKIP_HEADERS:
            continue
        field = COLUMN_MAP.get(normalized)
        if field:
            field_indexes[field] = index

    if "name" not in field_indexes:
        return [], ["Could not find a 'Company Name' column in the spreadsheet."]

    clients: list[dict[str, str | None]] = []
    errors: list[str] = []

    for row_number, row in enumerate(data_rows, start=header_index + 2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        client: dict[str, str | None] = {
            "name": None,
            "tan": None,
            "pan": None,
            "person_name": None,
            "gstin": None,
            "phone": None,
            "pr_phone": None,
            "pr_mobile": None,
        }

        for field, index in field_indexes.items():
            value = row[index] if index < len(row) else None
            client[field] = _cell_value(value)

        if not client["name"]:
            errors.append(f"Row {row_number}: skipped because Company Name is empty.")
            continue

        clients.append(client)

    return clients, errors
